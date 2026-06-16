from __future__ import annotations

import re
import shutil
import tempfile
import uuid
import zipfile
from datetime import datetime
from pathlib import Path

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from starlette.requests import Request

from gerar_automacao_extratos import (
    OUT_DIR,
    classify,
    extract_accounts,
    extract_transactions,
)


app = FastAPI(title="Desenvolva Lançamentos API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

BASE = Path(__file__).resolve().parent
CONTAS = BASE / "Contas.xls"
MODELO_DOMINIO = BASE / "modelo_dominio.xlsm"
JOBS_DIR = OUT_DIR / "api_jobs"
JOBS_DIR.mkdir(parents=True, exist_ok=True)

HISTORICOS = {
    "PAGAMENTO SIMPLES NACIONAL": (61, None),
    "ADIANTAMENTO DE LUCROS JEFFERSON": (62, None),
    "RECEBIMENTO COOPERATIVA": (24, None),
    "PAGAMENTO COMBUSTIVEL": (15, "COMBUSTIVEL"),
}


def slug(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "arquivo"


def period_label(period: dict) -> str:
    return f"{period['mes']:02d}_{period['ano']}"


def build_conference_xlsx(entries: list[dict], transactions: list[dict], output: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Lancamentos"
    headers = [
        "Data",
        "Debitar",
        "Conta Debitar",
        "Creditar",
        "Conta Creditar",
        "Valor",
        "Historico",
        "Descricao",
        "Regra",
        "Status",
    ]
    ws.append(headers)
    for entry in entries:
        ws.append(
            [
                entry["data"],
                entry["debitar_codigo"],
                entry["debitar_nome"],
                entry["creditar_codigo"],
                entry["creditar_nome"],
                entry["valor"],
                entry["historico"],
                entry["descricao_extrato"],
                entry["regra"],
                entry["conferido"],
            ]
        )

    bruto = wb.create_sheet("Extrato Bruto")
    bruto.append(["Data lancamento", "Data contabil", "Tipo", "Descricao", "Valor"])
    for tx in transactions:
        bruto.append(
            [
                tx["data_lancamento"],
                tx["data_contabil"],
                tx["tipo"],
                tx["descricao"],
                tx["valor"],
            ]
        )

    for sheet in wb.worksheets:
        if sheet.max_row > 1:
            table = Table(
                displayName=f"tbl{slug(sheet.title)}",
                ref=f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}",
            )
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            width = min(48, max(12, max(len(str(c.value or "")) + 2 for c in sheet[letter])))
            sheet.column_dimensions[letter].width = width
        sheet.freeze_panes = "A2"

    for sheet_name in ["Lancamentos", "Extrato Bruto"]:
        ws2 = wb[sheet_name]
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = "dd/mm/yyyy"
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

    wb.save(output)


def build_dominio_xlsm(entries: list[dict], output: Path) -> None:
    workbook = load_workbook(MODELO_DOMINIO, keep_vba=True)
    ws = workbook["Plan1"]
    ws["G3"] = str(output.parent)

    for row in range(6, max(ws.max_row, len(entries) + 6) + 1):
        for col in range(1, 11):
            ws.cell(row=row, column=col).value = None

    for idx, entry in enumerate(entries, start=6):
        hist_code, complemento = HISTORICOS.get(entry["historico"], ("", entry["historico"]))
        values = [
            entry["data"],
            entry["debitar_codigo"],
            entry["creditar_codigo"],
            entry["valor"],
            hist_code,
            complemento,
            123 if idx == 6 else None,
            44,
            None,
            None,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col)
            cell.value = value
            if col == 1:
                cell.number_format = "dd/mm/yyyy"
            elif col == 4:
                cell.number_format = "#,##0.00"

    workbook.save(output)


def build_entradas_txt(entries: list[dict], output: Path) -> None:
    lines = []
    for idx, entry in enumerate(entries):
        hist_code, complemento = HISTORICOS.get(entry["historico"], ("", entry["historico"]))
        fields = [
            entry["data"].strftime("%d/%m/%Y"),
            str(entry["debitar_codigo"] or ""),
            str(entry["creditar_codigo"] or ""),
            f"{float(entry['valor']):.2f}".replace(".", ","),
            str(hist_code or ""),
            str(complemento or ""),
            "123" if idx == 0 else "",
            "44",
            "",
            "",
        ]
        lines.append(";".join(fields))
    output.write_text("\n".join(lines) + "\n", encoding="latin1")


def process_statement(pdf_path: Path, password: str, job_dir: Path) -> dict:
    accounts = extract_accounts(CONTAS)
    accounts_by_code = {account["codigo"]: account for account in accounts}
    transactions, period = extract_transactions(pdf_path, password)
    entries = [classify(tx, accounts_by_code) for tx in transactions]

    label = period_label(period)
    conference = job_dir / f"conferencia_{label}.xlsx"
    dominio = job_dir / f"dominio_lancamentos_{label}.xlsm"
    entradas = job_dir / f"entradas_{label}.txt"
    pacote = job_dir / f"desenvolva_lancamentos_{label}.zip"

    build_conference_xlsx(entries, transactions, conference)
    build_dominio_xlsm(entries, dominio)
    build_entradas_txt(entries, entradas)

    with zipfile.ZipFile(pacote, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(conference, conference.name)
        zf.write(dominio, dominio.name)
        zf.write(entradas, entradas.name)

    total_entradas = sum(tx["valor"] for tx in transactions if tx["tipo"].startswith("Entrada"))
    total_saidas = sum(tx["valor"] for tx in transactions if tx["tipo"].startswith("Sa"))
    qtd_revisar = sum(1 for entry in entries if entry["conferido"] == "REVISAR")

    return {
        "competencia": label,
        "qtd_movimentos": len(transactions),
        "qtd_revisar": qtd_revisar,
        "total_entradas": round(total_entradas, 2),
        "total_saidas": round(total_saidas, 2),
        "total_liquido": round(total_entradas - total_saidas, 2),
        "lancamentos": [
            {
                "data": entry["data"].strftime("%Y-%m-%d"),
                "descricao": entry["descricao_extrato"],
                "debito": entry["debitar_codigo"],
                "credito": entry["creditar_codigo"],
                "valor": entry["valor"],
                "historico": entry["historico"],
                "status": entry["conferido"],
            }
            for entry in entries
        ],
        "arquivos": {
            "conferencia": conference.name,
            "modelo_dominio": dominio.name,
            "entradas_txt": entradas.name,
            "pacote_zip": pacote.name,
        },
    }


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "produto": "Desenvolva Lançamentos"}


@app.post("/processar-extrato")
async def processar_extrato(
    request: Request,
    arquivo: UploadFile = File(...),
    senha: str = Form(""),
) -> dict:
    if not arquivo.filename:
        raise HTTPException(status_code=400, detail="Arquivo do extrato nao enviado.")

    job_id = uuid.uuid4().hex[:12]
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    upload_name = slug(Path(arquivo.filename).stem) + Path(arquivo.filename).suffix.lower()
    uploaded = job_dir / upload_name
    temp_path = None

    try:
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            shutil.copyfileobj(arquivo.file, tmp)
            temp_path = Path(tmp.name)
        shutil.move(str(temp_path), uploaded)
        temp_path = None

        result = process_statement(uploaded, senha, job_dir)
    except Exception as exc:
        if temp_path and temp_path.exists():
            temp_path.unlink()
        raise HTTPException(
            status_code=500,
            detail={"erro": "Falha ao processar extrato.", "detalhe": str(exc)},
        ) from exc
    finally:
        await arquivo.close()

    base_url = str(request.base_url).rstrip("/")
    result["job_id"] = job_id
    result["downloads"] = {
        key: f"{base_url}/download/{job_id}/{filename}"
        for key, filename in result["arquivos"].items()
    }
    return result


@app.get("/download/{job_id}/{arquivo}")
def download(job_id: str, arquivo: str) -> FileResponse:
    jobs_root = JOBS_DIR.resolve()
    job_dir = (jobs_root / job_id).resolve()
    file_path = (job_dir / arquivo).resolve()

    try:
        job_dir.relative_to(jobs_root)
        file_path.relative_to(job_dir)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado.") from exc

    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Arquivo nao encontrado.")

    return FileResponse(
        file_path,
        media_type="application/octet-stream",
        filename=file_path.name,
    )
