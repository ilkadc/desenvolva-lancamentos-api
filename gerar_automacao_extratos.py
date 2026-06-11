from __future__ import annotations

import argparse
import json
import re
import struct
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


BASE = Path(__file__).resolve().parent
XLS_CONTAS = BASE / "Contas.xls"
PDF_EXTRATO = BASE / "extrato_03_2026.pdf"
OUT_DIR = BASE / "outputs"


def _read_compound_workbook_stream(path: Path) -> bytes:
    data = path.read_bytes()

    def u32(offset: int) -> int:
        return struct.unpack_from("<I", data, offset)[0]

    def u16(offset: int) -> int:
        return struct.unpack_from("<H", data, offset)[0]

    sector_size = 1 << u16(30)
    fat_count = u32(44)
    first_dir_sector = u32(48)
    difat = [
        sid
        for sid in struct.unpack_from("<109I", data, 76)
        if sid not in (0xFFFFFFFF, 0xFFFFFFFE)
    ]

    def sector_offset(sector_id: int) -> int:
        return (sector_id + 1) * sector_size

    fat: list[int] = []
    for sector_id in difat[:fat_count]:
        fat.extend(
            struct.unpack_from(
                f"<{sector_size // 4}I", data, sector_offset(sector_id)
            )
        )

    def read_chain(start_sector: int) -> bytes:
        output = bytearray()
        sector_id = start_sector
        seen: set[int] = set()
        while (
            sector_id not in (0xFFFFFFFE, 0xFFFFFFFF)
            and sector_id < len(fat)
            and sector_id not in seen
        ):
            seen.add(sector_id)
            offset = sector_offset(sector_id)
            output.extend(data[offset : offset + sector_size])
            sector_id = fat[sector_id]
        return bytes(output)

    directory = read_chain(first_dir_sector)
    for offset in range(0, len(directory), 128):
        entry = directory[offset : offset + 128]
        if len(entry) < 128:
            break
        name_len = struct.unpack_from("<H", entry, 64)[0]
        if name_len < 2:
            continue
        name = entry[: name_len - 2].decode("utf-16le", errors="ignore")
        start = struct.unpack_from("<I", entry, 116)[0]
        size = struct.unpack_from("<Q", entry, 120)[0]
        if name in {"Workbook", "Book"}:
            return read_chain(start)[:size]
    raise RuntimeError("Nao foi possivel localizar a aba Workbook no arquivo XLS.")


def extract_accounts(path: Path) -> list[dict]:
    data = _read_compound_workbook_stream(path)
    records: list[tuple[int, bytes]] = []
    pos = 0
    while pos + 4 <= len(data):
        record_type, length = struct.unpack_from("<HH", data, pos)
        pos += 4
        records.append((record_type, data[pos : pos + length]))
        pos += length

    sst_parts: list[bytes] = []
    in_sst = False
    for record_type, payload in records:
        if record_type == 0x00FC:
            sst_parts = [payload]
            in_sst = True
        elif in_sst and record_type == 0x003C:
            sst_parts.append(payload)
        elif in_sst:
            break

    sst = b"".join(sst_parts)
    strings: list[str] = []
    if sst:
        unique_count = struct.unpack_from("<I", sst, 4)[0]
        offset = 8
        for _ in range(unique_count):
            char_count = struct.unpack_from("<H", sst, offset)[0]
            options = sst[offset + 2]
            offset += 3
            has_rich_text = options & 0x08
            has_ext = options & 0x04
            is_utf16 = options & 0x01
            rich_runs = 0
            ext_size = 0
            if has_rich_text:
                rich_runs = struct.unpack_from("<H", sst, offset)[0]
                offset += 2
            if has_ext:
                ext_size = struct.unpack_from("<I", sst, offset)[0]
                offset += 4
            byte_count = char_count * (2 if is_utf16 else 1)
            raw = sst[offset : offset + byte_count]
            offset += byte_count
            strings.append(
                raw.decode("utf-16le" if is_utf16 else "latin1", errors="ignore")
            )
            if has_rich_text:
                offset += 4 * rich_runs
            if has_ext:
                offset += ext_size

    cells: dict[tuple[int, int], object] = {}
    sheet = 0
    for record_type, payload in records:
        if record_type == 0x0809:
            bof_type = struct.unpack_from("<H", payload, 2)[0] if len(payload) >= 4 else None
            if bof_type == 0x0010:
                sheet += 1
        elif record_type == 0x00FD and len(payload) >= 10:
            row, col, _xf, sst_index = struct.unpack_from("<HHHI", payload, 0)
            if sheet == 1:
                cells[(row, col)] = strings[sst_index] if sst_index < len(strings) else ""
        elif record_type == 0x0203 and len(payload) >= 14:
            row, col, _xf = struct.unpack_from("<HHH", payload, 0)
            if sheet == 1:
                value = struct.unpack_from("<d", payload, 6)[0]
                cells[(row, col)] = int(value) if value.is_integer() else value

    accounts: list[dict] = []
    for row in range(5, 2000):
        code = cells.get((row, 0), "")
        classification = cells.get((row, 7), "")
        account_type = cells.get((row, 3), "")
        name = ""
        for col in range(11, 18):
            value = cells.get((row, col), "")
            if (
                isinstance(value, str)
                and value.strip()
                and not re.fullmatch(r"[0-9.]+", value.strip())
            ):
                name = value.strip()
                break
        if code != "" and classification != "" and name:
            accounts.append(
                {
                    "codigo": int(code) if isinstance(code, (int, float)) else code,
                    "classificacao": str(classification),
                    "tipo": str(account_type),
                    "nome": name,
                }
            )
    return accounts


def extract_transactions(path: Path, password: str) -> tuple[list[dict], dict]:
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        reader.decrypt(password)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    period_match = re.search(
        r"Período • 1 de (?P<mes>[a-zç]+) de (?P<ano>\d{4}) até", text, re.IGNORECASE
    )
    month_names = {
        "janeiro": 1,
        "fevereiro": 2,
        "março": 3,
        "marco": 3,
        "abril": 4,
        "maio": 5,
        "junho": 6,
        "julho": 7,
        "agosto": 8,
        "setembro": 9,
        "outubro": 10,
        "novembro": 11,
        "dezembro": 12,
    }
    year = int(period_match.group("ano")) if period_match else 2026
    month_name = period_match.group("mes").lower() if period_match else ""
    month = month_names.get(month_name, 1)

    pattern = re.compile(
        r"(?P<data_lanc>\d{2}/\d{2})\s+"
        r"(?P<data_cont>\d{2}/\d{2})\s+"
        r"(?P<tipo>Entrada|Saída)\s+PIX\s+"
        r"(?P<descricao>.+?)\s+"
        r"(?P<valor>-?R\$\s*[\d.]+,\d{2})"
    )
    transactions = []
    for match in pattern.finditer(text):
        raw_value = match.group("valor").replace("R$", "").replace("\xa0", " ").strip()
        amount = float(raw_value.replace(" ", "").replace(".", "").replace(",", "."))
        transactions.append(
            {
                "data_lancamento": datetime.strptime(
                    f"{match.group('data_lanc')}/{year}", "%d/%m/%Y"
                ),
                "data_contabil": datetime.strptime(
                    f"{match.group('data_cont')}/{year}", "%d/%m/%Y"
                ),
                "tipo": match.group("tipo"),
                "descricao": match.group("descricao").strip(),
                "valor": abs(amount),
            }
        )
    return transactions, {"ano": year, "mes": month, "mes_nome": month_name}


def account_by_code(accounts: list[dict], code: int) -> dict:
    for account in accounts:
        if account["codigo"] == code:
            return account
    raise KeyError(code)


def classify(transaction: dict, accounts_by_code: dict[int, dict]) -> dict:
    desc = transaction["descricao"].upper()
    bank = accounts_by_code[8]

    if transaction["tipo"] == "Entrada" and "COOPERATIVA" in desc:
        debit = bank
        credit = accounts_by_code[519]
        hist = "RECEBIMENTO COOPERATIVA"
        rule = "Entrada PIX Cooperativa -> baixa cliente"
    elif transaction["tipo"] == "Saída" and "JEFFERSON ADEMIR" in desc:
        debit = accounts_by_code[520]
        credit = bank
        hist = "ADIANTAMENTO DE LUCROS JEFFERSON"
        rule = "Saida PIX Jefferson -> adiantamento de lucros"
    elif transaction["tipo"] == "Saída" and "RECEITA FEDERAL" in desc:
        debit = accounts_by_code[479]
        credit = bank
        hist = "PAGAMENTO SIMPLES NACIONAL"
        rule = "Saida PIX Receita Federal -> Simples Nacional a recolher"
    elif transaction["tipo"] == "Saída" and "COMBUST" in desc:
        debit = accounts_by_code[364]
        credit = bank
        hist = "PAGAMENTO COMBUSTIVEL"
        rule = "Saida PIX combustivel -> despesa combustivel"
    else:
        debit = {}
        credit = bank if transaction["tipo"] == "Saída" else {}
        hist = "REVISAR LANCAMENTO"
        rule = "Sem regra automatica"

    return {
        "data": transaction["data_contabil"],
        "tipo_lancamento": "Um débito para um crédito",
        "debitar_codigo": debit.get("codigo", ""),
        "debitar_classificacao": debit.get("classificacao", ""),
        "debitar_nome": debit.get("nome", ""),
        "creditar_codigo": credit.get("codigo", ""),
        "creditar_classificacao": credit.get("classificacao", ""),
        "creditar_nome": credit.get("nome", ""),
        "valor": transaction["valor"],
        "historico": hist,
        "descricao_extrato": transaction["descricao"],
        "regra": rule,
        "conferido": "OK" if debit and credit else "REVISAR",
    }


def autosize(ws, min_width: int = 10, max_width: int = 46) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = min_width
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def add_table(ws, name: str) -> None:
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def build_workbook(accounts: list[dict], transactions: list[dict], period: dict, output_file: Path) -> None:
    accounts_by_code = {account["codigo"]: account for account in accounts}
    entries = [classify(tx, accounts_by_code) for tx in transactions]

    wb = Workbook()
    ws = wb.active
    ws.title = "Lancamentos Dominio"
    headers = [
        "Data",
        "Tipo",
        "Debitar",
        "Classif. Debitar",
        "Nome Debitar",
        "Creditar",
        "Classif. Creditar",
        "Nome Creditar",
        "Valor",
        "Historico",
        "Descricao Extrato",
        "Regra Aplicada",
        "Conferido",
    ]
    ws.append(headers)
    for entry in entries:
        ws.append(
            [
                entry["data"],
                entry["tipo_lancamento"],
                entry["debitar_codigo"],
                entry["debitar_classificacao"],
                entry["debitar_nome"],
                entry["creditar_codigo"],
                entry["creditar_classificacao"],
                entry["creditar_nome"],
                entry["valor"],
                entry["historico"],
                entry["descricao_extrato"],
                entry["regra"],
                entry["conferido"],
            ]
        )
    for cell in ws["A"][1:]:
        cell.number_format = "dd/mm/yyyy"
    for cell in ws["I"][1:]:
        cell.number_format = '#,##0.00'
    add_table(ws, "tblLancamentosDominio")
    style_sheet(ws)

    resumo = wb.create_sheet("Resumo")
    resumo.append(["Item", "Valor"])
    resumo.append(["Quantidade de movimentos", len(entries)])
    resumo.append(["Total entradas", sum(tx["valor"] for tx in transactions if tx["tipo"] == "Entrada")])
    resumo.append(["Total saidas", sum(tx["valor"] for tx in transactions if tx["tipo"] == "Saída")])
    resumo.append(["Saldo liquido do extrato", "=B3-B4"])
    resumo.append(["Movimentos para revisar", '=COUNTIF(\'Lancamentos Dominio\'!M:M,"REVISAR")'])
    for cell in resumo["B"][2:5]:
        cell.number_format = '#,##0.00'
    add_table(resumo, "tblResumo")
    style_sheet(resumo)

    regras = wb.create_sheet("Regras")
    regras.append(
        [
            "Prioridade",
            "Tipo extrato",
            "Texto procurado",
            "Debitar",
            "Debitar nome",
            "Creditar",
            "Creditar nome",
            "Historico",
            "Observacao",
        ]
    )
    rules = [
        (1, "Entrada", "COOPERATIVA", 8, "BANCO C6 BANK", 519, "COOPERATIVA DE TRABALHO...", "RECEBIMENTO COOPERATIVA", "Confirme se deve baixar cliente ou reconhecer receita."),
        (2, "Saída", "JEFFERSON ADEMIR", 520, "ADIANTAMENTO DE LUCROS", 8, "BANCO C6 BANK", "ADIANTAMENTO DE LUCROS JEFFERSON", ""),
        (3, "Saída", "RECEITA FEDERAL", 479, "SIMPLES NACIONAL A RECOLHER", 8, "BANCO C6 BANK", "PAGAMENTO SIMPLES NACIONAL", ""),
        (4, "Saída", "COMBUST", 364, "COMBUSTIVEL", 8, "BANCO C6 BANK", "PAGAMENTO COMBUSTIVEL", "Troque para conta 292 se o gasto for custo direto."),
    ]
    for row in rules:
        regras.append(row)
    add_table(regras, "tblRegras")
    style_sheet(regras)

    bruto = wb.create_sheet("Extrato Bruto")
    bruto.append(["Data lancamento", "Data contabil", "Tipo", "Descricao", "Valor"])
    for tx in transactions:
        bruto.append(
            [
                tx["data_lancamento"],
                tx["data_contabil"],
                tx["tipo"],
                tx["descricao"],
                tx["valor"],
            ]
        )
    for col in ("A", "B"):
        for cell in bruto[col][1:]:
            cell.number_format = "dd/mm/yyyy"
    for cell in bruto["E"][1:]:
        cell.number_format = '#,##0.00'
    add_table(bruto, "tblExtratoBruto")
    style_sheet(bruto)

    plano = wb.create_sheet("Plano de Contas")
    plano.append(["Codigo", "Classificacao", "Tipo/Sintetica", "Nome"])
    for account in accounts:
        plano.append(
            [
                account["codigo"],
                account["classificacao"],
                account["tipo"],
                account["nome"],
            ]
        )
    add_table(plano, "tblPlanoContas")
    style_sheet(plano)

    OUT_DIR.mkdir(exist_ok=True)
    wb.save(output_file)

    diagnostics = {
        "arquivo": str(output_file),
        "movimentos": len(transactions),
        "entradas": round(sum(tx["valor"] for tx in transactions if tx["tipo"] == "Entrada"), 2),
        "saidas": round(sum(tx["valor"] for tx in transactions if tx["tipo"] == "Saída"), 2),
        "revisar": sum(1 for entry in entries if entry["conferido"] == "REVISAR"),
    }
    diag_name = f"diagnostico_automacao_{period['mes']:02d}_{period['ano']}.json"
    (OUT_DIR / diag_name).write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", type=Path, default=PDF_EXTRATO)
    parser.add_argument("--senha", default="583051")
    parser.add_argument("--saida", type=Path)
    args = parser.parse_args()

    contas = extract_accounts(XLS_CONTAS)
    extrato, periodo = extract_transactions(args.pdf, args.senha)
    saida = args.saida or OUT_DIR / f"automacao_lancamentos_extrato_{periodo['mes']:02d}_{periodo['ano']}.xlsx"
    build_workbook(contas, extrato, periodo, saida)
